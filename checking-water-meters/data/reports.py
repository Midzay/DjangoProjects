from django.shortcuts import render 
from django.contrib.auth.models import User
from .models import Main, City, DefaultForUser, View_svidet,PreviouslyFormedProtocol, Etalon
from django.views.generic.edit import CreateView, View
from django.http.response import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import copier
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from .functions import *
import csv


def one_report(request):
    l = []
    nm = []

    for f in Main._meta.get_fields():
        if Main._meta.get_field(f.name).verbose_name in request.POST:
            l.append(f.name)
            nm.append(Main._meta.get_field(f.name).verbose_name)
    
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORT'
    ws.merge_cells('B1:E1')
    j = 2
    # заголовок таблицы
    ws.cell(row=3, column=j - 1).value = '#'
    for el in nm:
        if el == 'Серия свидетельства':
            continue
        ws.cell(row=3, column=j).value = el
        j += 1
    row = ws.row_dimensions[3]
    row.font = Font(bold=True)
    row.height = 40
    row.alignment = Alignment(horizontal="center", wrap_text=True)
    request.session['date_s'] = request.POST['date_s']
    request.session['date_po'] = request.POST['date_po']
    request.session['search_adr'] = request.POST['search_adr']
    request.session['view_svidet'] = request.POST['view_svidet']
    request.session['city'] = request.POST['city']
    request.session['water'] = request.POST['water']
    request.session['num_fabric'] = request.POST['num_fabric']
    request.session['num_svidet_powerka'] = request.POST['num_svidet_powerka']
    request.session['fio_worker'] = request.POST['fio_worker']

    s = request.session['date_s']
    po = request.session['date_po']
    adr = request.session['search_adr']
    view_svidet = request.session['view_svidet']
    c = request.session['city']
    wat = request.session['water']

    q = create_request_to_bd(s=s, po=po, adr=adr)
    if view_svidet != '0':
        q = q & Q(view_svidet__view_svidet=view_svidet)

    if wat != '9':
        q = q & Q(hot_water=wat)
    if c != '0':
        q = q & Q(city__city=c)
    main = Main.objects.filter(q)
    i = 4
    count_row = 1
    # заполняем значениями
    for row in main:

        j = 2
        ws.cell(row=i, column=j - 1).value = count_row
        for cel in l:
            if 'date' in cel:
                result = '.'.join(str(getattr(row, cel)).split('-')[::-1])
            else:
                result = str(getattr(row, cel))
            if cel == 'result_powerka':
                if str(getattr(row, cel)) == '1':
                    result = "Пригодно"
                else:
                    result = 'Непригодно'
            if cel == 'view_svidet':

                if getattr(row, cel) == None:
                    result = "-"
                else:
                    result = str(getattr(row, cel))
            if cel == 'hot_water':
                if str(getattr(row, cel)) == '1':
                    result = "Гор."
                else:
                    result = 'Хол.'
            if cel == 'plomba':
                if str(getattr(row, cel)) == '1':
                    result = "Да"
                else:
                    result = 'Нет'
            if cel == 'view_svidet':
                var_viw_swidet = str(getattr(row, cel))
                continue

            if cel == 'num_svidet_powerka':
                result = var_viw_swidet + ' ' + str(getattr(row, cel))

            ws.cell(row=i, column=j).value = result
            ws.column_dimensions[get_column_letter(j)].width = 15
            j += 1
        count_row += 1
        row = ws.row_dimensions[i]
        row.alignment = Alignment(
            horizontal="center", wrap_text=True, shrink_to_fit=False, )
        i += 1
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename=report.xlsx'
    response['Content-Disposition'] = content
    wb.save(response)
    return response


class MiniReport(View):
    obj_report = ''

    def add_lst(self, lst, name, all_cnt, agr):
        if agr['cnt'] > 0:
            proverka_ok = agr['sum_p']
            proverka_no = all_cnt['cnt'] - agr['sum_p']
            hot_water = agr['sum_w']
            cold_water = agr['cnt'] - agr['sum_w']
            lst.append([name, proverka_ok, proverka_no, hot_water, cold_water])
        else:
            lst.append([name, 0, 0, 0, 0])
        return lst

    def if_q(self, q, s, po):
        if s != '':
            q = q & Q(date__gte=s)
        if po != '':
            q = q & Q(date__lte=po)
        return q

    def get(self, request):
        context = {}
        lst = []
        if self.obj_report == 'Work':
            lst = Worker().get_workers(request)
            lst.sort(key=lambda l: l[1], reverse=True)
            context['name_table'] = 'ФИО Работника'
        elif self.obj_report == 'City':
            lst = City_report().get_city(request)
            context['name_table'] = 'Город'
            lst.sort(key=lambda l: l[1], reverse=True)

        context['fio'] = lst

        agr = Main.objects.all().aggregate(
            cnt=Count('date'), sum_p=Sum('result_powerka'))
        if agr['cnt'] > 0:
            context['all'] = agr['cnt']
            context['pow_ok'] = agr['sum_p']
            context['pow_no'] = agr['cnt'] - agr['sum_p']
            # print (context)
        return render(request, 'worker_report.html', context=context)

    def post(self, request):
        context = {}
        lst = []
        if self.obj_report == 'Work':
            lst, context = Worker().post_workers(request)
            lst.sort(key=lambda l: l[1], reverse=True)
            context['name_table'] = 'ФИО Работника'
        elif self.obj_report == 'City':
            lst, context = City_report().post_city(request)
            lst.sort(key=lambda l: l[1], reverse=True)
            context['name_table'] = 'Город'

        # print(lst), 7777
        context['fio'] = lst
        agr = Main.objects.all().aggregate(
            cnt=Count('date'), sum_p=Sum('result_powerka'))
        if agr['cnt'] > 0:
            context['all'] = agr['cnt']
            context['pow_ok'] = agr['sum_p']
            context['pow_no'] = agr['cnt'] - agr['sum_p']

        return render(request, 'worker_report.html', context=context)


class Worker(MiniReport):
    def get_workers(self, request):
        lst = []
        for usr in User.objects.all():
            q = Q(fio_worker__exact=usr.last_name)
            all_cnt = Main.objects.filter(
                fio_worker__exact=usr.last_name).aggregate(cnt=Count('date'))
            agr = Main.objects.filter(q).aggregate(
                cnt=Count('date'), sum_w=Sum('hot_water'), sum_p=Sum('result_powerka'))
            name = usr.last_name
            lst = self.add_lst(lst, name, all_cnt, agr)
        #   print (agr)

        return lst

    def post_workers(self, request):
        context = {}
        lst = []
        for usr in User.objects.all():
            q = Q(fio_worker__exact=usr.last_name)
            q = self.if_q(q, request.POST['date_s'], request.POST['date_po'])
            context['date_s'] = request.POST['date_s']
            context['date_po'] = request.POST['date_po']
            all_cnt = Main.objects.filter(q).aggregate(cnt=Count('date'))
            agr = Main.objects.filter(q).aggregate(
                cnt=Count('date'), sum_w=Sum('hot_water'), sum_p=Sum('result_powerka'))

            name = usr.last_name
            lst = self.add_lst(lst, name, all_cnt, agr)
        return lst, context


class City_report(MiniReport):
    def get_city(self, request):
        lst = []
        for city in City.objects.all():
            q = Q(city__city__exact=city.city)
            all_cnt = Main.objects.filter(
                city__city__exact=city.city).aggregate(cnt=Count('date'))
            agr = Main.objects.filter(q).aggregate(
                cnt=Count('date'), sum_w=Sum('hot_water'), sum_p=Sum('result_powerka'))
            name = city
            lst = self.add_lst(lst, name, all_cnt, agr)
        #    print(lst)
        return lst

    def post_city(self, request):
        context = {}
        lst = []
        for city in City.objects.all():
            q = Q(city__city__exact=city.city)
            q = self.if_q(q, request.POST['date_s'], request.POST['date_po'])
            context['date_s'] = request.POST['date_s']
            context['date_po'] = request.POST['date_po']
            all_cnt = Main.objects.filter(q).aggregate(cnt=Count('date'))
            agr = Main.objects.filter(q).aggregate(
                cnt=Count('date'), sum_w=Sum('hot_water'), sum_p=Sum('result_powerka'))
            name = city
            lst = self.add_lst(lst, name, all_cnt, agr)
        return lst, context


def protocol(request, num_id):

    try:
        rec1 = Main.objects.get(id=num_id)
        def_user = DefaultForUser.objects.get(
            user=User.objects.get(last_name=rec1.fio_worker).id)
    except:
        pass
    try:
        protokol13 = def_user.protokol13
    except:
        protokol13 = 0
    try:
        protokol14 = def_user.protokol14
    except:
        protokol14 = 0
    try:
        protokol15 = def_user.protokol15
    except:
        protokol15 = 0

    try:
        name_for_protokol = def_user.name_for_protokol
    except:
        name_for_protokol = 0

    if Main.objects.filter(id=num_id).count() != 0:
        rec = Main.objects.get(id=num_id)
    dt_sled = int(str(rec.date_end_powerka).split('-')[0])
    dt = int(str(rec.date).split('-')[0])
    interval = dt_sled - dt
    if interval < 5:
        txt = str(interval) + " года"
    else:
        txt = str(interval) + " лет"
    # wb = load_workbook('powerka/data/templates/data/tempate.xlsx')
    wb = load_workbook('data/templates/data/tempate.xlsx')  # dev

    ws = wb.active
    ws['B4'] = str(rec.view_svidet) + ' ' + str(rec.num_svidet_powerka)
    if rec.hot_water == 1:
        ws['B5'] = "счетчик горячей воды крыльчатый"
    else:
        ws['B5'] = "счетчик холодной воды крыльчатый"
    ws['B6'] = rec.name_type
    ws['D6'] = rec.num_fabric
    ws['B7'] = rec.num_gos_reestr
    ws['E7'] = txt
    ws['B8'] = rec.fio_client
    ws['B9'] = 'г. ' + str(rec.city) + ', ' + rec.adress
    ws['B10'] = str(rec.method)
    etalon = Etalon.objects.get(etalon = rec.etalon)
    ws['B12'] = str(etalon.full_name)
    if protokol13:
        ws['B13'] = protokol13
    if protokol14:
        ws['B14'] = protokol14
    if protokol15:
        ws['B15'] = protokol15

    ws['D38'] = rec.qh
    ws['D40'] = str(rec.view_svidet) + ' ' + str(rec.num_svidet_powerka)
    ws['E40'] = 'от ' + '.'.join(str(rec.date).split('-')[::-1])
    ws['B41'] = '.'.join(str(rec.date).split('-')[::-1])
    if name_for_protokol:
        ws['D42'] = name_for_protokol
    else:
        ws['D42'] = request.user.first_name
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename=protocol.xlsx'
    response['Content-Disposition'] = content
    wb.save(response)

    return response


def reportForFif(request):
    request.session['date_s'] = request.POST['date_s']
    request.session['date_po'] = request.POST['date_po']
    request.session['search_adr'] = request.POST['search_adr']
    request.session['view_svidet'] = request.POST['view_svidet']
    request.session['city'] = request.POST['city']
    request.session['water'] = request.POST['water']

    s = request.session['date_s']
    po = request.session['date_po']
    adr = request.session['search_adr']
    view_svidet = request.session['view_svidet']
    c = request.session['city']
    wat = request.session['water']

    q = create_request_to_bd(s=s, po=po, adr=adr)
    if view_svidet != '0':
        q = q & Q(view_svidet__view_svidet=view_svidet)

    if wat != '9':
        q = q & Q(hot_water=wat)
    if c != '0':
        q = q & Q(city__city=c)
    q = q & Q(result_powerka=1)
    main = Main.objects.filter(q)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reportFIF.csv";encoding="utf-8"'
    response.write(u'\ufeff'.encode('utf8'))
    filewriter = csv.writer(response, delimiter=';',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    filewriter.writerow(
        ['TypePOV', 'GosNumberPOV', 'NamePOV', 'DesignationSiPOV', 'DeviceMarkPOV', 'DeviceCountPOV',
         'SerialNumPOV', 'SerialNumEndPOV', 'CalibrationDatePOV', 'NextcheckDatePOV', 'MarkCipherPOV', 'DocPOV',
         'DeprcatedPOV', 'NumCertfPOV', 'NumSvidPOV', 'PrimPOV', 'ScopePOV', 'StandartPOV', 'GpsPOV', 'SiPOV',
         'SoPOV'])
    for row in main:
        filewriter.writerow(
            [1, row.num_gos_reestr, '', '', '', 1, row.num_fabric, '', row.date, row.date_end_powerka, '',
             str(row.method),
             'Пригодно', '', str(row.view_svidet) + ' ' + str(row.num_svidet_powerka), '', '', row.etalon, '', '', ''])

    return response

# Блок проверки данных и свидетельств. Печать протоколов


def check_data(request):
    context = {'svidet_object': View_svidet.objects.all()}
    context['fio_worker_list'] = User.objects.all()
    context['previously_formed'] = PreviouslyFormedProtocol.objects.all().order_by('-id')[:10]
    return render(request, 'superuser/check.html', context=context)


def check_number_svidet(request):
    try:
        num_start = int(request.POST['num_start'])
        num_end = int(request.POST['num_end'])
        view_svidet = request.POST['view_svidet']

        range_svidetelstv = range(num_start, num_end+1)

        # Передаем  серию свидетельста и диапазон. должны получить номера которых нет и их кол-во.
        wb = Workbook()
        ws = wb.active

        # заголовок таблицы
        j = 2
        ws.cell(row=2, column=j).value = 'Cерия свидетельства'
        ws.cell(row=2, column=j+2).value = view_svidet
        
        ws.cell(row=3, column=j).value = 'Кол-во:'

        ws.cell(row=4, column=j).value = 'Номер свидетельства '
        row = ws.row_dimensions[3]
        row.font = Font(bold=True)
        row.height = 40
        row.alignment = Alignment(horizontal="center", wrap_text=True)

        main = Main.objects.filter(view_svidet__view_svidet=view_svidet).filter(
            num_svidet_powerka__gte=num_start).filter(num_svidet_powerka__lte=num_end)

        if not main:
            context ={}
            context = {'svidet_object': View_svidet.objects.all()}
            context['fio_worker_list'] = User.objects.all()
            context['previously_formed'] = PreviouslyFormedProtocol.objects.all().order_by('-id')[:10]
            context['error']=1
            return render(request, 'superuser/check.html', context=context)
            
        number_in_base = set(map(int,[el['num_svidet_powerka'] for el in list(main.values())]))
        
        miss_number = []
        for el in range_svidetelstv:
            if el not in number_in_base:
                miss_number.append(el)
        
        ws.cell(row=3, column=j+2).value = str(len(miss_number))
        # заполняем значениями
        j=2
        i=5
        for num in  miss_number:
            
            ws.cell(row=i, column=j).value = str(num)
            ws.column_dimensions[get_column_letter(j)].width = 15
            
            
            row = ws.row_dimensions[i]
            row.alignment = Alignment(
                horizontal="center", wrap_text=True, shrink_to_fit=False, )
            i+=1
        
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename=mis_number.xlsx'
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except:
        context ={}
        context = {'svidet_object': View_svidet.objects.all()}
        context['fio_worker_list'] = User.objects.all()
        context['previously_formed'] = PreviouslyFormedProtocol.objects.all().order_by('-id')[:10]
        context['error']=1
        return render(request, 'superuser/check.html', context=context)


from copy import copy
import random
import datetime

def random_table (flag_date,paskal):
    temp_water = round(random.uniform(9.5,16.0),1)
    temp = round(random.uniform(19.0,25.0),1)
    vlagnost = round(random.uniform(30.0,59.0),1)
    if flag_date:
        paskal =  round(random.uniform(748.0,752.0),1)
    return temp_water, temp, vlagnost, paskal

# Печать многих протоколов
def many_protokol(request):
    
    try:
        num_start = int(request.POST['num_start'])
        num_end = int(request.POST['num_end'])
        view_svidet = request.POST['view_svidet']
        fio_worker = request.POST['fio_worker']
        
        b = PreviouslyFormedProtocol(
            fio_worker=fio_worker,
            view_svidet=view_svidet,
            range_number = str(num_start)+'--'+str(num_end),
            )
        b.save()
    except: pass

    protocols =[]
    try:
        protocols = Main.objects.filter(view_svidet__view_svidet=view_svidet).filter(
        num_svidet_powerka__gte=num_start).filter(num_svidet_powerka__lte=num_end).filter(fio_worker=fio_worker)
        

        def_user = DefaultForUser.objects.get(
            user=User.objects.get(last_name=fio_worker).id)
    except:
        pass
    try:
        protokol13 = def_user.protokol13
    except:
        protokol13 = 0
    try:
        protokol14 = def_user.protokol14
    except:
        protokol14 = 0
    try:
        protokol15 = def_user.protokol15
    except:
        protokol15 = 0

    try:
        name_for_protokol = def_user.name_for_protokol
    except:
        name_for_protokol = 0

    # wb = load_workbook('powerka/data/templates/data/tempate.xlsx')
    wb = load_workbook('data/templates/data/tempate.xlsx')  # dev
    
    data_main = datetime.datetime(2000, 1, 1, 0, 0, 0)
    paskal =  round(random.uniform(748.0,752.0),1)
    if protocols:
        for i,rec in enumerate(protocols):
            ws = wb.worksheets[i]
            ws.title = f'Протокол {i+1}'
            
            dt_sled = int(str(rec.date_end_powerka).split('-')[0])
            dt = int(str(rec.date).split('-')[0])
            interval = dt_sled - dt
        
            flag_date = 0 
            if data_main != rec.date:
                data_main =rec.date
                flag_date = 1


            #Заполлняем мальенькую таблицу
            temp_water, temp, vlagnost, paskal = random_table (flag_date, paskal)
            ws['E18'], ws['F18'] = temp_water,temp_water
            ws['E19'], ws['F19'] = temp,temp
            ws['E20'], ws['F20'] = vlagnost, vlagnost
            ws['E21'], ws['F21'] = paskal, paskal

            print (temp_water, temp, vlagnost, paskal)

            if interval < 5:
                txt = str(interval) + " года"
            else:
                txt = str(interval) + " лет"
            
            ws['B4'] = str(rec.view_svidet) + ' ' + str(rec.num_svidet_powerka)
            if rec.hot_water == 1:
                ws['B5'] = "счетчик горячей воды крыльчатый"
            else:
                ws['B5'] = "счетчик холодной воды крыльчатый"
            ws['B6'] = rec.name_type
            ws['D6'] = rec.num_fabric
            ws['B7'] = rec.num_gos_reestr
            ws['E7'] = txt
            ws['B8'] = rec.fio_client
            ws['B9'] = 'г. ' + str(rec.city) + ', ' + rec.adress
            ws['B10'] = str(rec.method)
            etalon = Etalon.objects.get(etalon = rec.etalon)
            ws['B12'] = str(etalon.full_name)
            if protokol13:
                ws['B13'] = protokol13
            if protokol14:
                ws['B14'] = protokol14
            if protokol15:
                ws['B15'] = protokol15

            ws['D38'] = rec.qh
            ws['D40'] = str(rec.view_svidet) + ' ' + str(rec.num_svidet_powerka)
            ws['E40'] = 'от ' + '.'.join(str(rec.date).split('-')[::-1])
            ws['B41'] = '.'.join(str(rec.date).split('-')[::-1])
            if name_for_protokol:
                ws['D42'] = name_for_protokol
            else:
                ws['D42'] = request.user.first_name
            # создаем новый лист и копирум все данные и стили
            ws_new = wb.create_sheet('del',i+1)
            c = copier.WorksheetCopy(ws, ws_new)
            c.copy_worksheet()
        del_sheet=wb.get_sheet_by_name('del')
        wb.remove_sheet(del_sheet)
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename=protocol.xlsx'
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    context ={}
    context = {'svidet_object': View_svidet.objects.all()}
    context['fio_worker_list'] = User.objects.all()
    context['previously_formed'] = PreviouslyFormedProtocol.objects.all().order_by('-id')[:10]
    context['error']=1
    return render(request, 'superuser/check.html', context=context)
    
