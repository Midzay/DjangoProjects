from django.shortcuts import render, HttpResponse
from django.views.generic.edit import CreateView, View
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from .classforusers import *
from datetime import datetime
import xlrd
from datetime import timedelta
from django.utils import timezone

from dateutil.relativedelta import relativedelta

import calendar



# Create your views here.


class Index(View):

    def get(self, request):
        if request.user.is_authenticated:
            if request.user.groups.filter(name='dispatcher').exists():
                main = Dispatcher.objects.all()
                context = ClassPaginator().pagin(request, main)
                context['form'] = MainForm
                #print(context, 78787878787)
                return render(request, 'dispatcher/superuser/index.html', context)
        else:
            return redirect('/accounts/login/')

    def post(self, request):
        if request.user.is_authenticated:
            if request.user.groups.filter(name='dispatcher').exists():
                bound_form = MainForm(request.POST, )
                if bound_form.is_valid():
                    bound_form.save()
                    return redirect('/dispatcher/')
                return render(request, 'dispatcher/superuser/index.html', context={'form': bound_form})
        else:
            return redirect('/accounts/login/')


class Edit(View):

    def get(self, request, num_id):
        con = Dispatcher.objects.get(id=num_id)
        bound_form = MainForm(instance=con)
        return render(request, 'dispatcher/superuser/edit_index.html', context={'form': bound_form, 'con': con})

    def post(self, request, num_id):
        con = Dispatcher.objects.get(id=num_id)
        bound_form = MainForm(request.POST, instance=con)
        if bound_form.is_valid():
            bound_form.save()
            return redirect('/data/')
        return render(request, 'edit.html', context={'form': bound_form, 'con': con})


def del_dispatcher(request, num_id):
    main = Dispatcher.objects.all()
    if Dispatcher.objects.filter(id=num_id).count() != 0:
        rec = Dispatcher.objects.get(id=num_id)
        rec.delete()
        return redirect('/dispatcher/')
    else:
        return render(request, 'dispatcher/superuser/index.html', context={'main': main})


class RevizorClass():
    def create_context(self, request):

        if request.method == 'GET':
            q = Q()
            main = Dispatcher.objects.filter(q)
            context = ClassPaginator().pagin(request, main)
            context['form'] = MainForm

            return (request, 'dispatcher/superuser/disp_reports.html', context)

        else:

            q = Q()
            if request.POST['date'] != '':
                q = q | Q(date=request.POST['date'])
            if request.POST['fio_client'] != '':
                q = q | Q(fio_client=request.POST['fio_client'])
            if request.POST['phone'] != '':
                q = q | Q(phone=request.POST['phone'])
            if request.POST['adress'] != '':
                q = q | Q(adress=request.POST['adress'])
            if request.POST['name_type'] != '':
                q = q | Q(name_type=request.POST['name_type'])
            if request.POST['num_fabric'] != '':
                q = q | Q(num_fabric=request.POST['num_fabric'])
            if request.POST['pokazaniya'] != '':
                q = q | Q(pokazaniya=request.POST['pokazaniya'])
            if request.POST['last_update'] != '':
                q = q | Q(last_update=request.POST['last_update'])
            if request.POST['hot_water'] != '9':
                q = q & Q(hot_water=request.POST['hot_water'])
            if request.POST['num_dogovor'] != '':
                q = q | Q(num_dogovor=request.POST['num_dogovor'])
            if request.POST['empty_pokaz'] != '':
                date_befor_month = datetime.today() - relativedelta(months=int(request.POST['empty_pokaz']))
                #print (str(date_befor_month).split()[0],989898989)
                q = q & Q(last_update__lte=str(date_befor_month).split()[0])

            main = Dispatcher.objects.filter(q)
            context = ClassPaginator().pagin(request, main)
            context['date'] = request.POST['date']
            context['fio_client'] = request.POST['fio_client']
            context['phone'] = request.POST['phone']
            context['adress'] = request.POST['adress']
            context['name_type'] = request.POST['name_type']
            context['num_fabric'] = request.POST['num_fabric']
            context['pokazaniya'] = request.POST['pokazaniya']
            context['last_update'] = request.POST['last_update']
            context['hot_water'] = request.POST['hot_water']
            context['num_dogovor'] = request.POST['num_dogovor']
            context['empty_pokaz'] = request.POST['empty_pokaz']
            #context['form'] = MainForm(request.POST)

            # print('!!!!!!!!!!',context)
            return (request, 'dispatcher/superuser/disp_reports.html', context)


def one_report(request):
    l = []
    nm = []

    for f in Dispatcher._meta.get_fields():
        if Dispatcher._meta.get_field(f.name).verbose_name in request.POST:
            l.append(f.name)
            nm.append(Dispatcher._meta.get_field(f.name).verbose_name)

    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'Отчет по показаниям на '+str(datetime.today().strftime("%d-%m-%Y")) +':'
    ws.merge_cells('B1:E1')
    j = 2
    # заголовок таблицы
    ws.cell(row=3, column=j - 1).value = '#'
    for el in nm:
        ws.cell(row=3, column=j).value = el
        j += 1
    row = ws.row_dimensions[3]
    row.font = Font(bold=True)
    row.height = 40
    row.alignment = Alignment(horizontal="center", wrap_text=True)
    q = Q()
    if request.POST['date'] != '':
        q = q | Q(date=request.POST['date'])
    if request.POST['fio_client'] != '':
        q = q | Q(fio_client=request.POST['fio_client'])
    if request.POST['phone'] != '':
        q = q | Q(phone=request.POST['phone'])
    if request.POST['adress'] != '':
        q = q | Q(adress=request.POST['adress'])
    if request.POST['name_type'] != '':
        q = q | Q(name_type=request.POST['name_type'])
    if request.POST['num_fabric'] != '':
        q = q | Q(num_fabric=request.POST['num_fabric'])
    if request.POST['pokazaniya'] != '':
        q = q | Q(pokazaniya=request.POST['pokazaniya'])
    if request.POST['last_update'] != '':
        q = q | Q(last_update=request.POST['last_update'])
    if request.POST['hot_water'] != '9':
        q = q & Q(hot_water=request.POST['hot_water'])
    if request.POST['num_dogovor'] != '':
        q = q | Q(num_dogovor=request.POST['num_dogovor'])
    if request.POST['empty_pokaz'] != '':
        date_befor_month = datetime.today() - relativedelta(months=int(request.POST['empty_pokaz']))
       # print(str(date_befor_month).split()[0], 989898989)
        q = q & Q(last_update__lte=str(date_befor_month).split()[0])


    main = Dispatcher.objects.filter(q)
    i = 4
    count_row = 1
    # заполняем значениями
    for row in main:
        j = 2
        ws.cell(row=i, column=j - 1).value = count_row
        for cel in l:
            if 'date' in cel:
                result = '.'.join(str(getattr(row, cel)).split('-')[::-1])
            else:
                result = str(getattr(row, cel))
            if cel == 'hot_water':
                if str(getattr(row, cel)) == '1':
                    result = "Гор."
                else:
                    result = 'Хол.'
            ws.cell(row=i, column=j).value = result
            ws.column_dimensions[get_column_letter(j)].width = 15
            j += 1
        count_row += 1
        row = ws.row_dimensions[i]
        row.alignment = Alignment(
            horizontal="center", wrap_text=True, shrink_to_fit=False, )
        i += 1
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename=report.xlsx'
    response['Content-Disposition'] = content
    wb.save(response)
    return response


def report_for_dispatcher(request):
    if request.method == "POST" and 'btn_rep' in request.POST:
        return one_report(request)
    else:
        whoiam = RevizorClass()
        r, t, c = whoiam.create_context(request)
        return render(r, t, c)

def update(request):
    if request.method == "GET":

        return render(request, 'dispatcher/superuser/update.html', context=dict())
    else:
        # print(request.FILES)
        excel_file = request.FILES["excel_file"]
        # print (excel_file,676776767)
        # wb = load_workbook(excel_file)
        # worksheet = wb.sheetnames
        # print(worksheet)
        wb =xlrd.open_workbook(file_contents=excel_file.read())
        xl_sheet =wb.sheet_by_index(0)
        for row_idx in range(2, xl_sheet.nrows):
            eh_value =str(int(xl_sheet.cell(row_idx,1).value))
            pokaz =str((xl_sheet.cell(row_idx,2).value))
            last_up = str(xl_sheet.cell(row_idx,3).value.split()[0]).split('.')
            last_up = '-'.join([last_up[2],last_up[1],last_up[0]])
           # print (data.split()[0])

            Dispatcher.objects.filter(num_fabric=eh_value).update(pokazaniya=pokaz,last_update=last_up)
        return redirect('/dispatcher/')





#
# def main_form(request):
#     if request.user.is_authenticated:
#         if request.user.groups.filter(name='dispatcher').exists():
#             whoiam = SuperUserClass()
#             r, s, c = whoiam.create_context(request)
#             return render(r, s, c)
#     else:
#         return redirect('/accounts/login/')
#
# def add_rec(request):
#     bound_form = MainForm(request.POST,)
#     if bound_form.is_valid():
#         bound_form.save()
#         return render(request, 'dispatcher/superuser/index.html', context={'form': bound_form, })
#     return render(request, 'dispatcher/superuser/index.html', context={'form': bound_form, })
